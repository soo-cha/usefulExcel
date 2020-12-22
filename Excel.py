'''
최초 파일 수행시
pip install openpyxl 수행하여 필수 library를 다운받아야 한다.
'''

import openpyxl                         # download한 openpyxl을 import 한다.


#wb = openpyxl.Workbook()                # workbook은 엑셀 파일 자체를 의미한다. (새로운 엑셀 파일 만들기)
wb = openpyxl.load_workbook('./msa.xlsx')  # 기존의 엑셀 파일을 불러오는 방식


#sheet = wb.active                       # 만들어진 wb 객체를 active하면 sheet 객체를 만들수 있다.
#sheet2 = wb.create_sheet('sheet2')      # 두번째 sheet를 만들수 있다.

sheet = wb['Sheet1']                     # 만약 두번째 시트가 이미 존재하는 파일이라면 sheet 이름을 통해 가져올 수 있다.
#sheet.title = '1반 출석부'              # sheet의 이름 바꾸기


'''
for i in range(5):
    sheet.cell(row=1, column=i+1).value = '김민아'    #반복문을 활용한 5개 줄에 이름 넣기    
'''


# 모든 행 단위로 출력 (출석부 excel에서 학생들 이름을 가져와서 성적 Excel에 자동 세팅이 가능해짐)
print('모든 행 출력')
rowNum = 1
for row in sheet.rows:
    print(str(rowNum)+'행 시작-------------')
    for cell in row:
        print(cell.value)
    rowNum = rowNum +1
    
print('모든 행 출력 종료')

'''
# 셀좌표값에 접근 후 출력 
print('cell(2,2) 값은?=============='+sheet.cell(2,2).value) 
# 셀 범위 지정하여 차례대로 출력
get_cells = sheet['A1':'C3'] 
for row in get_cells:
    for cell in row: 
        print(cell.value)
'''



wb.save('test.xlsx')                    # 엑셀 파일을 저장한다.